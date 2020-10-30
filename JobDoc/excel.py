from openpyxl import load_workbook

dockerFolder = "file/"

def addJob(fileName:str,data:dict):
	filename = dockerFolder+fileName+".xlsx"
	workbook = load_workbook(filename)
	sheet = workbook.active
	lastRow = sheet.max_row


	if lastRow == 1:
		sheet.cell(row=lastRow,column=1).value = "Name:"
		sheet.cell(row=lastRow,column=2).value = "Date:"
		sheet.cell(row=lastRow,column=3).value = "Importance:"
		sheet.cell(row=lastRow,column=4).value = "Languages:"
		sheet.cell(row=lastRow,column=5).value = "Website:"


	name:str = data["name"]
	date:str = str(data["date"].month)+"-"+str(data["date"].day)+"-"+str(data["date"].year)
	importance:str = str(data["importance"])
	languages:str = data["languages"]
	website:str = data["website"]

	sheet.cell(row=lastRow+1,column=1).value = name
	sheet.cell(row=lastRow+1,column=2).value = date
	sheet.cell(row=lastRow+1,column=3).value = importance
	sheet.cell(row=lastRow+1,column=4).value = languages
	sheet.cell(row=lastRow+1,column=5).value = website

	workbook.save(filename)


def sortExcel(filename:str,criteria:str):
	filename = dockerFolder+filename+".xlsx"
	workbook = load_workbook(filename)
	sheet = workbook.active
	lastRow = sheet.max_row
	colNum = 0

	if criteria == "name":
		colNum = 1

	elif criteria == "date":
		colNum = 2

	elif criteria == "importance":
		colNum = 3

	quickSort(sheet,2,lastRow,colNum)

	workbook.save(filename)


def quickSort(excel:object,low:int,high:int,criteria:int):
	if low < high:
		pi = partition(excel,low,high,criteria)

		quickSort(excel,low,pi-1,criteria)
		quickSort(excel,pi+1,high,criteria)

def partition(excel:object,low:int,high:int,criteria:int):
	i = low -1

	for j in range(low, high):
		if getValue(excel,j,criteria) >= getValue(excel,high,criteria) and criteria == 3:
			i = i+1
			swapRows(excel,i,j)

		elif getValue(excel,j,criteria) <= getValue(excel,high,criteria) and (criteria == 1 or criteria == 2):
			i = i+1
			swapRows(excel,i,j)

	swapRows(excel,i+1,high)
	return i+1


def swapRows(excel:object,firstRow:int,secondRow:int):
	fRowName = getValue(excel,firstRow,1)
	fRowDate = getValue(excel,firstRow,2)
	fRowImportance = getValue(excel,firstRow,3)
	fRowLanguages = getValue(excel,firstRow,4)
	fRowWebsite = getValue(excel,firstRow,5)

	sRowName = getValue(excel,secondRow,1)
	sRowDate = getValue(excel,secondRow,2)
	sRowImportance = getValue(excel,secondRow,3)
	sRowLanguages = getValue(excel,secondRow,4)
	sRowWebsite = getValue(excel,secondRow,5)

	setValue(excel,firstRow,1,sRowName)
	setValue(excel,firstRow,2,sRowDate)
	setValue(excel,firstRow,3,sRowImportance)
	setValue(excel,firstRow,4,sRowLanguages)
	setValue(excel,firstRow,5,sRowWebsite)

	setValue(excel,secondRow,1,fRowName)
	setValue(excel,secondRow,2,fRowDate)
	setValue(excel,secondRow,3,fRowImportance)
	setValue(excel,secondRow,4,fRowLanguages)
	setValue(excel,secondRow,5,fRowWebsite)



def getValue(excel:object,rowVal:int,colVal:int):
	return excel.cell(row=rowVal,column=colVal).value

def setValue(excel:object,rowVal:int,colVal:int,newVal:object):
	excel.cell(row=rowVal,column=colVal).value = newVal


