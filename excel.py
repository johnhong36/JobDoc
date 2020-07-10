from openpyxl import load_workbook

def addJob(fileName:str,data:dict):
	workbook = load_workbook(filename=fileName+".xlsx")
	sheet = workbook.active
	lastRow = sheet.max_row


	if lastRow == 1:
		sheet.cell(row=lastRow,column=1).value = "Name:"
		sheet.cell(row=lastRow,column=2).value = "Date:"
		sheet.cell(row=lastRow,column=3).value = "Importance:"
		sheet.cell(row=lastRow,column=4).value = "Languages:"
		sheet.cell(row=lastRow,column=5).value = "Website:"


	name:str = data["name"]
	date:str = data["date"]
	importance:str = data["importance"]
	languages:str = data["languages"]
	website:str = data["website"]

	sheet.cell(row=lastRow+1,column=1).value = name
	sheet.cell(row=lastRow+1,column=2).value = date
	sheet.cell(row=lastRow+1,column=3).value = importance
	sheet.cell(row=lastRow+1,column=4).value = languages
	sheet.cell(row=lastRow+1,column=5).value = website

	workbook.save(filename=fileName+".xlsx")

# if __name__ == "__main__":
# 	ex = {
# 		"name": "SRC",
# 		"date": "3/6/99",
# 		"importance": "4",
# 		"languages": "Python",
# 		"website": ""
# 	}
# 	addJob("hello_world",ex)
